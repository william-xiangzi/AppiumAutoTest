#!/usr/bin/python
# -*- coding: UTF-8 -*-

from openpyxl import Workbook
from openpyxl import load_workbook


class Caseread(object):
    try:
        casedata = load_workbook('./AppiumCase/autotestcase-backup.xlsx')
    except (AttributeError, FileNotFoundError):
        print("读取文件出错")

    # 注释：用字典的方法存储Appium Case；
    # 注释：取值用法 appiumcase 中 key 值是用例名称；有几个key 即有几条 case 用法：len(appiumcase)
    # 注释：每个 case 中的操作步骤使用 二维list 存值；二维list 长度为几 case actionnumber 即为几 用法：len(Caseread.readdata()['匹配音频'])
    # 注释：每个 actionnumber list 中：0 -- Elementlocatemode、1 -- Element、2 -- Action、3 -- Content
    @classmethod
    def readdata(cls):
        casenumber = len(cls.casedata.sheetnames)
        casename = cls.casedata.sheetnames
        # 用例值存字典；读取到之后直接写入字典；即可存储多个值
        appiumcase = {}
        for index in range(casenumber):
            case = cls.casedata[casename[index]]
            row = 3
            actionum = case.cell(row=row, column=2).value
            casedetail = []
            while actionum:
                elementname = case.cell(row=row, column=3).value
                elementlocatemode = case.cell(row=row, column=4).value
                element = case.cell(row=row, column=5).value
                action = case.cell(row=row, column=6).value
                content = case.cell(row=row, column=7).value
                attribute = case.cell(row=row, column=8).value
                casedetail.append([elementlocatemode, element, action, content, elementname, attribute])
                appiumcase[casename[index]] = casedetail
                row += 1
                actionum = case.cell(row=row, column=2).value
        return appiumcase


if __name__ == '__main__':
    print(Caseread.readdata())
    print(len(Caseread.readdata()))
    print(len(Caseread.readdata()['用户登录']))
